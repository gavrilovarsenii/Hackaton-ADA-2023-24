import logging, os
import nest_asyncio
import pandas as pd
import numpy as np
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ApplicationBuilder, ContextTypes, CommandHandler, MessageHandler, filters, CallbackQueryHandler
import openpyxl
# Assuming you have a 'keys.py' file with your token
import keys

# Apply nest_asyncio to allow nested event loops
nest_asyncio.apply()

df = pd.read_excel("db.xlsx",header=1)
df = df.iloc[: , 1:]
df.columns = df.columns.astype(str)
df.columns = [x[:20] for x in df.columns]

def load_workbook_and_sheet():
    global wb, sheet
    wb = openpyxl.load_workbook(data_base_filename)
    sheet = wb.active

data_base_filename = 'db.xlsx'

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logging.getLogger("httpx").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# Load your Excel workbook
load_workbook_and_sheet()

check_admin = False

# Function to check if the user is the bot administrator
def is_admin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    user = update.effective_user
    return user.id == keys.admin_id

async def upload_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if check_admin and not is_admin(update, context):
        await update.message.reply_text("You're not authorized to use this command.")
        return

    await update.message.reply_text("Please enter the password:")
    context.user_data['waiting_for_password'] = True
    return

async def cancel_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['waiting_for_file'] = False
    context.user_data['waiting_for_password'] = False
    await update.message.reply_text("The uploading canceled")

async def downloader(update, context):
    if context.user_data['waiting_for_password']:
        await update.message.reply_text("Before uploading please enter the password (Use /cancel_upload to cancel the upload if you want to): ")
        return

    if not context.user_data['waiting_for_file']:
        await update.message.reply_text("Please use /upload_excel command to upload files")
        return

    # Download file
    fileName = update.message.document.file_name
    new_file = await update.message.effective_attachment.get_file()
    await new_file.download_to_drive(fileName)
    os.replace(fileName, data_base_filename)

    context.user_data['waiting_for_file'] = False

    # Update the file
    load_workbook_and_sheet()

    # Acknowledge file received
    await update.message.reply_text(f"{fileName} saved successfully")
    return

# Function to read column names from the second column onwards
def get_column_names():
    return [cell.value for cell in sheet[2][1:]]

# Command handler for /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text('Введите номер региона для получения информации о налоговом вычете (номер региона может отличаться от реального)')
    context.user_data['waiting_for_password'] = False
    context.user_data['waiting_for_file'] = False

# Message handler for word input
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get('waiting_for_password'):
        try:
            if update.message.text == str(keys.password):
                context.user_data['waiting_for_file'] = True
                context.user_data['waiting_for_password'] = False
                await update.message.reply_text("Password accepted. Please upload the Excel file.")
            else:
                await update.message.reply_text("Incorrect password. Access denied.")
                context.user_data['waiting_for_password'] = False
        except Exception as e:
            logging.error(f"Error in handle_password: {e}")
            context.user_data['waiting_for_password'] = False
        return

    if context.user_data['waiting_for_file']:
        await update.message.reply_text("Please upload the Excel file. Use /cancel_upload to cancel the upload")
        return


    index = str(update.message.text)
    dataarr = list(df.iloc[int(index)])
    if(index.isdigit() and int(index) < df.shape[0]):
        # Create an inline keyboard with column names
        keyboard = [
        [InlineKeyboardButton(str(column_name)[:20], callback_data=f"{index}:{str(column_name)[:20]}")]
        for column_name in df.columns[1:] if not(pd.isna(df[column_name][int(index)]))]
        keyboard.append([InlineKeyboardButton("Выбрать другой регион",callback_data="back")])
        #print([strcolumn_name  for column_name in df.columns[1:] if not(pd.isna(df[column_name][int(index)]))])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(f'Ваш выбор {df[df.columns[0]][int(index)]}! Доступная информация:', reply_markup=reply_markup)
        return
    await update.message.reply_text('Данный номер региона не найден в текущей базе данных')

# Callback query handler for column selection
async def handle_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    query.answer()
    if query.data == "back":
      await query.edit_message_text(f"Введите номер региона")
      return
    index, column_name = query.data.split(':')
    # Find the row with the word and get the value from the selected column
    keyboard = [
    [InlineKeyboardButton(str(column_name)[:20], callback_data=f"{index}:{str(column_name)[:20]}")]
    for column_name in df.columns[1:] if not(pd.isna(df[column_name][int(index)]))]
    keyboard.append([InlineKeyboardButton("Выбрать другой регион",callback_data="back")])
    keyboard.remove([InlineKeyboardButton(str(column_name)[:20], callback_data=f"{index}:{str(column_name)[:20]}")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(f"Ваш выбор {df[df.columns[0]][int(index)]}! По вопросу {column_name} имеются следующие данные: {df[column_name][int(index)]}", reply_markup=reply_markup) #(f'The value for "{word}" in column "{column_name}" is: {value}')
    #await update.message.reply_text('новый тест:', reply_markup=reply_markup)
    return

if __name__ == '__main__':
    application = ApplicationBuilder().token(keys.token).build()

    # Command handler for /start
    start_handler = CommandHandler('start', start)
    application.add_handler(start_handler)

    # Message handler for word input
    message_handler = MessageHandler(filters.TEXT & ~filters.COMMAND & ~filters.ATTACHMENT, handle_message)
    application.add_handler(message_handler)

    upload_excel_handler = CommandHandler('upload_excel', upload_excel)
    application.add_handler(upload_excel_handler)

    upload_excel_handler = CommandHandler('cancel_upload', cancel_upload)
    application.add_handler(upload_excel_handler)

    application.add_handler(MessageHandler(filters.ATTACHMENT, downloader))

    # Callback query handler for column selection
    button_handler = CallbackQueryHandler(handle_button)
    application.add_handler(button_handler)

    # Run the bot
    application.run_polling()
